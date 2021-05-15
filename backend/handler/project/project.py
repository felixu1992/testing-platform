import json
import os
import time

from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from openpyxl import load_workbook
from rest_framework import serializers, viewsets
from rest_framework.decorators import action

from backend.exception import ErrorCode, ValidateError, PlatformError
from backend.handler.case import case_info
from backend.handler.project import project_group
from backend.handler.record import record
from backend.models import Project, CaseInfo, Report
from backend.settings import IGNORED, FAILED, PASSED
from backend.util import UserHolder, Response, parse_data, get_params, update_fields, page_params, Executor, save, \
    batch_save
from backend.util.resp_data import obj_to_dict
from testing_platform.settings import FILE_REPO


class ProjectSerializer(serializers.ModelSerializer):
    class Meta:
        model = Project
        fields = ['id', 'name', 'remark', 'headers', 'host', 'group_id', 'notify', 'created_at', 'updated_at']


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects

    serializer_class = ProjectSerializer

    def create(self, request, *args, **kwargs):
        """
        创建项目
        """

        body = parse_data(request, 'POST')
        project = Project(**body)
        # 校验分组是否存在
        project_group.get_by_id(project.group_id)
        save(project)
        return Response.success(project)

    def destroy(self, request, *args, **kwargs):
        """
        根据 id 删除项目

        前端需要二次确认，确认删除执行以下操作：
        1. 删除项目信息
        2. 删除关联接口用例信息
        3. 删除对应项目的历史报告(或者不删，让他有个后悔机会还能找回接口)
        """

        parse_data(request, 'DELETE')
        id = kwargs['pk']
        project = get_by_id(id)
        project.delete()
        case_info.delete_by_project(id)
        return Response.def_success()

    def update(self, request, *args, **kwargs):
        """
        修改项目信息
        """

        data = parse_data(request, 'PUT')
        params = get_params(data, 'id', 'name', 'remark', 'headers', 'host', 'notify', 'group_id', toleration=True)
        project = get_by_id(params['id'])
        update_fields(project, **params)
        # 校验分组是否存在
        project_group.get_by_id(project.group_id)
        save(project)
        return Response.success(project)

    def list(self, request, *args, **kwargs):
        """
        分页查询项目

        可全量分页(当然只有自己的数据)
        可传入分组
        可传入 name 模糊
        """

        data = parse_data(request, 'GET')
        page, page_size, name, group_id = page_params(data, 'name', 'group_id').values()
        projects = Project.objects.filter(owner=UserHolder.current_user()).exact(group_id=group_id).contains(name=name)
        page_projects = Paginator(projects, page_size)
        result = page_projects.page(page)
        # 得到所有分组 id
        group_ids = [o.group_id for o in result.object_list]
        group_names = {}
        groups = project_group.get_list_by_ids(group_ids)
        for group in groups:
            group_names.update({group.id: group.name})
        for project in result.object_list:
            group_name = group_names[project.group_id]
            if group_name:
                project.group_name = group_name
        return Response.success(result)

    def retrieve(self, request, *args, **kwargs):
        """
        根据 id 查询项目详细信息
        """

        parse_data(request, 'GET')
        id = kwargs['pk']
        return Response.success(get_by_id(id))

    @action(methods=['POST'], detail=False, url_path='copy')
    def copy(self, request):
        """
        拷贝项目

        会拷贝以下内容：
        1. 项目信息
        2. 项目下接口信息
        """

        data = parse_data(request, 'POST')
        id, name = get_params(data, 'id', 'name').values()
        old_project = get_by_id(id)
        new_project = Project()
        new_project.__dict__ = old_project.__dict__.copy()
        new_project.id = None
        new_project.name = name
        save(new_project)
        case_infos = case_info.list_by_project(old_project.id)
        new_infos = []
        for info in case_infos:
            new_case_info = CaseInfo()
            new_case_info.__dict__ = info.__dict__.copy()
            new_case_info.id = None
            new_case_info.project_id = new_project.id
            new_infos.append(new_case_info)
        if new_infos:
            batch_save(CaseInfo.objects, new_infos)
        return Response.success(new_project)

    @action(methods=['POST'], detail=False, url_path='execute')
    def execute(self, request):
        """
        执行项目下所有接口用例

        1. 执行接口用例
        2. 生成用例报告
        """

        data = parse_data(request, 'POST')
        params = get_params(data, 'id')
        project = get_by_id(params['id'])
        case_infos = case_info.list_by_project(project.id)
        if not case_infos:
            raise PlatformError.error(ErrorCode.PROJECT_NOT_HAVE_CASES)
        executor = Executor(case_infos=case_infos, project=project)
        reports = executor.execute()
        # 构建记录
        ignored_num = [obj for obj in reports if getattr(obj, 'status') == IGNORED]
        passed_num = [obj for obj in reports if getattr(obj, 'status') == PASSED]
        failed_num = [obj for obj in reports if getattr(obj, 'status') == FAILED]
        reco = record.create(group_id=project.group_id, project_id=project.id, owner=project.owner,
                             total=len(case_infos), ignored=len(ignored_num), passed=len(passed_num),
                             failed=len(failed_num))
        for result in reports:
            result.record_id = reco.id
            case_info.encoding(result)
            result.owner = reco.owner
        batch_save(Report.objects, [Report(**obj_to_dict(report)) for report in reports])
        return Response.success(reco)

    @action(methods=['POST'], detail=False, url_path='temp-import')
    def imported(self, request):
        """
        旧版 Excel 的导入
        """
        data = parse_data(request, 'POST')
        params = get_params(data, 'project_id')
        project = get_by_id(params['project_id'])
        files = request.FILES.getlist('files')
        if len(files) == 0 or len(files) > 1:
            raise ValidateError.error(ErrorCode.VALIDATION_ERROR, '请检查文件数目，需要有且仅有一个文件')
        file = files[0]
        file_path = os.path.join(FILE_REPO, str(int(time.time() * 1000)) + file.name)
        # 写文件
        dest = open(file_path, 'wb+')
        for chunk in file.chunks():
            dest.write(chunk)
        dest.close()
        src_parser = ExcelParser(file_path)
        # 得到所有 sheet 页
        sheet_names = src_parser.get_sheet_names()
        # 循环 sheet 页
        cases = []
        for sheet_name in sheet_names:
            holder = CaseInfoHolder(src_parser.work_book, sheet_name)
            for case in holder.case_infos:
                info = CaseInfo()
                info.name = case.step
                info.remark = case.description
                info.method = case.method
                info.run = False if case.run else True
                info.host = None
                info.path = case.path
                info.headers = case.headers
                info.check_status = False
                info.delay = case.sleep if case.sleep else 0
                info.params = json.loads(case.params) if case.params else None
                info.sample = json.loads(case.response_content) if case.response_content and len(
                    case.response_content) < 30000 and case.response_content != 'None' else None
                info.project_id = project.id
                build_extend(info, case.ex_keys, case.ex_values, cases)
                build_expected(info, case.expected_key, case.expected_value, case.check_step)
                info.owner = project.owner
                case_info.create_case(info)
                info.row = case.row
                cases.append(info)
        if os.path.exists(file_path):
            os.remove(file_path)
        return Response.def_success()

    # @action(methods=['GET'], detail=False, url_path='temp-export')
    # def exported(self, request):
    #     """
    #     导出旧版 Excel
    #     """
    #     data = parse_data(request, 'POST')
    #     params = get_params(data, 'project_id')
    #     project = get_by_id(params['project_id'])
    #     case_infos = CaseInfo.objects.owner().exact(project_id=project.id)



    @action(methods=['GET'], detail=False, url_path='statistics')
    def statistics(self, request):
        """
        首页的统计数据
        """

        # 查询所有项目
        # 查询所有接口
        # 查询所有记录

        return Response.success({'project': Project.objects.owner().count(),
                                 'case': case_info.count(),
                                 'record': record.count()})


def build_extend(info, ex_keys, ex_values, cases):
    if not ex_keys:
        return
    keys = ex_keys.split(',')
    values = ex_values.split(',')
    extend_keys = []
    extend_values = []
    for i in range(0, len(keys)):
        extend_keys.append(keys[i].split('.'))
        value = values[i]
        depend_keys = value.split(':')
        row = int(depend_keys[0])
        filter_cases = [case for case in cases if case.row == row]
        depend = filter_cases[0].id
        steps = depend_keys[1].split('.')
        extend_values.append({
            'depend': depend,
            'steps': steps
        })
    info.extend_keys = extend_keys
    info.extend_values = extend_values


def build_expected(info, ex_keys, ex_values, check_steps):
    keys = ex_keys.split(',')
    ex_values = ex_values if isinstance(ex_values, str) else str(ex_values)
    values = ex_values.split(',')
    _steps = check_steps.split(',')
    expected_keys = []
    expected_values = []
    for i in range(0, len(keys)):
        expected_keys.append(keys[i].split('.'))
        value = values[i]
        steps = [{'value': step} for step in _steps[i].split('.')]
        expected_values.append({
            'depend': None,
            'steps': steps,
            'value': value
        })
    info.expected_keys = expected_keys
    info.expected_values = expected_values


class ExcelParser:

    def __init__(self, filename):
        """
        持有整个 Excel
        """
        self.work_book = load_workbook(filename)

    def get_sheet_names(self):
        """
        拿到当前 Excel 所有的 Sheet
        """
        return self.work_book.sheetnames

    def get_sheet_count(self):
        """
        拿到当前 Excel Sheet 的总数
        """
        return len(self.work_book.sheetnames)


class CaseInfoHolder:

    def __init__(self, work_book, sheet_name):
        """
        解析当前 sheet 页为 CaseInfo 对象列表

        :param work_book: 整个工作簿
        :param sheet_name: 当前 sheet 页
        """
        # 取工作簿中对应 sheet 页保存
        self.sheet = work_book[sheet_name]
        # 获取数据总行数
        self.rows = list(self.sheet.rows)
        # 获取表头列表
        self.title = [column.value for column in self.rows[4]]
        # 定义持有的用例列表
        self.case_infos = []
        # 获取登录行信息
        self.default_host = self.sheet.cell(row=4, column=2).value
        # 构建登录信息
        self.login_info = CaseInfo()
        self.build_longin_info()
        # 构建所有请求信息
        self.build_request_info()

    def build_longin_info(self):
        """
        构建请求登录的信息
        """
        self.login_info.method = 'post'
        self.login_info.host = self.sheet.cell(row=2, column=2).value
        params = self.sheet.cell(row=3, column=2).value
        self.login_info.params = {} if str_is_none(params) else json.loads(params)
        self.login_info.path = ''
        # 请求头为 application/x-www-form-urlencoded
        self.login_info.headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    def build_request_info(self):
        """
        构建所有用例信息
        """
        case_count = len(self.rows)
        # 从第六行开始取用例信息
        for row in range(5, case_count):
            # 获取整行信息为列表
            info = [column.value for column in self.rows[row]]
            # 定义当前要处理的用例对象
            case_info = CaseInfo()
            # 将表头字段对应的列作为对象字段，将当前行对应的列信息作为值，填入用例对象中
            for i in zip(self.title, info):
                setattr(case_info, i[0], i[1])
            # 定义当前用例所在行数(其实该行数为实际对应 Excel 中行数减 1)
            case_info.row = row
            # case_info.method = case_info.method.lower()
            # 如果期望值没填则忽略该数据
            expected_code = case_info.expected_key
            if expected_code is None or expected_code == '':
                continue
            # 加入用例列表
            self.case_infos.append(case_info)


# -------------------------------------------- 以上为 RESTFUL 接口，以下为调用接口 -----------------------------------------

def get_by_id(id):
    """
    根据 id 查询项目
    """

    try:
        project = Project.objects.get(owner=UserHolder.current_user(), id=id)
    except ObjectDoesNotExist:
        raise PlatformError.error_args(ErrorCode.DATA_NOT_EXISTED, '项目', 'id')
    return project


def get_list_by_ids(ids):
    """
    根据 id 数组查询
    """

    return Project.objects.owner().fields_in(id=ids)


def count_by_group(group_id):
    """
    统计指定 group id 下项目数
    """

    return Project.objects.filter(owner=UserHolder.current_user(), group_id=group_id).count()


# -------------------------------------------- 迁移代码 -----------------------------------------

def str_is_none(source):
    """
    判断字符串不为空
    """
    if source == '' or source == 'NULL' or source == 'None' or source is None:
        return True
    return False


def get_value(source, steps):
    """
    根据入参字典以及取值步骤取出结果

    取值步骤为 . 连接，如：data.records.0.name 含义为取 data 下 records 列表的第 0 条的 name 字段的值
    """
    # 分割取值步骤为列表
    keys = steps.split(".")
    try:
        # 循环取值步骤字典
        for i in range(0, len(keys)):
            # 取字段值
            key = keys[i]
            # 如果为数字则转为数字(数字代表从列表取值)，否则为字符
            key = key if not key.isdigit() else int(key)
            # 从结果字典取值
            source = source[key]
    # 出现异常直接填充为空字符
    except Exception:
        return ''
    return source


def set_value(value, target, steps):
    # 以 . 分割插入步骤为数组
    keys = steps.split(".")
    value_list = []
    # 循环找到对应位置插入
    for i in range(0, len(keys)):
        key = keys[i]
        # 如果当前步骤为字符串类型的数字，转为 int
        key = key if not key.isdigit() else int(key)
        # 取到最后了，直接将值插入该位置
        if i == len(keys) - 1:
            target[key] = value
        # 否则继续往下找插入位置
        # 如果末尾为数字，则认为入参为[]
        elif len(keys) > 1 and (keys[len(keys) - 1]).isdigit():
            value_list.append(value)
            target[keys[len(keys) - 2]] = value_list
            break
        else:
            try:
                # 从目标中取当前 key 对应的值为下一次的目标
                target = target[key]
            except KeyError:
                # 报错则下一次目标为空字典
                target.update({key: {}})
                target = target[key]
