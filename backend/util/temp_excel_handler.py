import json

import openpyxl
from openpyxl import load_workbook
from backend.models import CaseInfo


def build_extend(info, ex_keys, ex_values, cases):
    if not ex_keys:
        return
    keys = ex_keys.split(',')
    values = ex_values.split(',')
    extend_keys = []
    extend_values = []
    for i in range(0, len(keys)):
        extend_keys.append(keys[i].split('.'))
        value = values[i]
        depend_keys = value.split(':')
        row = int(depend_keys[0])
        filter_cases = [case for case in cases if case.row == row]
        depend = filter_cases[0].id
        steps = depend_keys[1].split('.')
        extend_values.append({
            'depend': depend,
            'steps': steps
        })
    info.extend_keys = extend_keys
    info.extend_values = extend_values


def parse_extend(case_info, case_infos):
    if not case_info or not case_info.extend_keys:
        return None, None
    case_info.extend_keys = json.loads(case_info.extend_keys)
    case_info.extend_values = json.loads(case_info.extend_values)
    keys = []
    for key in case_info.extend_keys:
        keys.append(key[0])
    # values = []
    # for value in case_info.extend_values:
    #     depends = [case for case in case_infos if case.id == int(value['depend'])]
    #     depend = depends[0]
    #     values.
    var = [str([case for case in case_infos if case.id == value['depend']][0].row) + ':' + ('.'.join(value['steps']))
           for value
           in case_info.extend_values]
    return ",".join(keys), ','.join(var)


def build_expected(info, ex_keys, ex_values, check_steps):
    keys = ex_keys.split(',')
    ex_values = ex_values if isinstance(ex_values, str) else str(ex_values)
    values = ex_values.split(',')
    _steps = check_steps.split(',')
    expected_keys = []
    expected_values = []
    for i in range(0, len(keys)):
        expected_keys.append(keys[i].split('.'))
        value = values[i]
        steps = [{'value': step} for step in _steps[i].split('.')]
        expected_values.append({
            'depend': None,
            'steps': steps,
            'value': value
        })
    info.expected_keys = expected_keys
    info.expected_values = expected_values


def parse_expected(case_info):
    expected_keys = []
    case_info.expected_keys = json.loads(case_info.expected_keys)
    case_info.expected_values = json.loads(case_info.expected_values)
    for key in case_info.expected_keys:
        expected_keys.append('.'.join(key))
    values = []
    steps = []
    for value in case_info.expected_values:
        values.append(value['value'])
        steps.append('.'.join([step['value'] for step in value['steps']]))
    return ','.join(expected_keys), ','.join(values), ','.join(steps)


class ExcelParser:

    def __init__(self, filename):
        """
        持有整个 Excel
        """
        self.work_book = load_workbook(filename)

    def get_sheet_names(self):
        """
        拿到当前 Excel 所有的 Sheet
        """
        return self.work_book.sheetnames

    def get_sheet_count(self):
        """
        拿到当前 Excel Sheet 的总数
        """
        return len(self.work_book.sheetnames)


class ExcelWriter:
    def __init__(self, filename, sheet_name):
        """
        初始化需要写入的 Excel
        """

        self.work_book = openpyxl.Workbook()
        self.sheet = self.work_book.active
        self.sheet.title = sheet_name
        self.filename = filename

    def write(self, case_infos):
        """
        开始写结果
        """

        # 写结果信息
        self.set_columns(case_infos)
        # 保存结果
        self.work_book.save(self.filename)
        # 关闭 Excel
        self.work_book.close()

    def set_columns(self, case_infos):
        """
        写结果信息
        """

        # 写表头
        headers = ['description', 'step', 'run', 'method', 'host', 'path', 'sleep', 'params', 'ex_keys', 'ex_values',
                   'headers', 'expected_key', 'expected_value', 'check_step']
        # 数据长度
        length = len(headers)
        for i in range(1, length):
            self.sheet.cell(row=1, column=i, value=headers[i - 1])
        # 遍历写所有用例信息
        for index in range(0, len(case_infos)):
            row = index + 2
            case_info = case_infos[index]
            case_info.row = row - 1
            self.sheet.cell(row=row, column=1, value=case_info.remark)
            self.sheet.cell(row=row, column=2, value=case_info.name)
            self.sheet.cell(row=row, column=3, value='no' if not case_info.run else None)
            self.sheet.cell(row=row, column=4, value=case_info.method)
            self.sheet.cell(row=row, column=5, value=case_info.host)
            self.sheet.cell(row=row, column=6, value=case_info.path)
            self.sheet.cell(row=row, column=7, value=case_info.delay if case_info.delay > 0 else None)
            self.sheet.cell(row=row, column=8, value=json.dumps(case_info.params, indent=2,
                                                                ensure_ascii=False) if case_info.params else None)
            ex_keys, ex_values = parse_extend(case_info, case_infos)
            self.sheet.cell(row=row, column=9, value=ex_keys)
            self.sheet.cell(row=row, column=10, value=ex_values)
            self.sheet.cell(row=row, column=11, value=case_info.headers)
            expected_key, expected_value, check_step = parse_expected(case_info)
            self.sheet.cell(row=row, column=12, value=expected_key)
            self.sheet.cell(row=row, column=13, value=expected_value)
            self.sheet.cell(row=row, column=14, value=check_step)


class CaseInfoHolder:

    def __init__(self, work_book, sheet_name):
        """
        解析当前 sheet 页为 CaseInfo 对象列表

        :param work_book: 整个工作簿
        :param sheet_name: 当前 sheet 页
        """
        # 取工作簿中对应 sheet 页保存
        self.sheet = work_book[sheet_name]
        # 获取数据总行数
        self.rows = list(self.sheet.rows)
        # 获取表头列表
        self.title = [column.value for column in self.rows[4]]
        # 定义持有的用例列表
        self.case_infos = []
        # 获取登录行信息
        self.default_host = self.sheet.cell(row=4, column=2).value
        # 构建登录信息
        self.login_info = CaseInfo()
        self.build_longin_info()
        # 构建所有请求信息
        self.build_request_info()

    def build_longin_info(self):
        """
        构建请求登录的信息
        """
        self.login_info.method = 'post'
        self.login_info.host = self.sheet.cell(row=2, column=2).value
        params = self.sheet.cell(row=3, column=2).value
        self.login_info.params = {} if str_is_none(params) else json.loads(params)
        self.login_info.path = ''
        # 请求头为 application/x-www-form-urlencoded
        self.login_info.headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    def build_request_info(self):
        """
        构建所有用例信息
        """
        case_count = len(self.rows)
        # 从第六行开始取用例信息
        for row in range(5, case_count):
            # 获取整行信息为列表
            info = [column.value for column in self.rows[row]]
            # 定义当前要处理的用例对象
            case_info = CaseInfo()
            # 将表头字段对应的列作为对象字段，将当前行对应的列信息作为值，填入用例对象中
            for i in zip(self.title, info):
                setattr(case_info, i[0], i[1])
            # 定义当前用例所在行数(其实该行数为实际对应 Excel 中行数减 1)
            case_info.row = row
            # case_info.method = case_info.method.lower()
            # 如果期望值没填则忽略该数据
            expected_code = case_info.expected_key
            if expected_code is None or expected_code == '':
                continue
            # 加入用例列表
            self.case_infos.append(case_info)


def str_is_none(source):
    """
    判断字符串不为空
    """
    if source == '' or source == 'NULL' or source == 'None' or source is None:
        return True
    return False
